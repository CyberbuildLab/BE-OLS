# File: json_to_excel.py

import json, sys
import pandas as pd
from openpyxl import load_workbook

def append_to_excel(json_path, excel_path, sheet_name='Data'):
    # 1. load existing sheet
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found")
    df_existing = pd.read_excel(excel_path, sheet_name=sheet_name)
    # 2. load new record
    with open(json_path, 'r', encoding='utf-8') as f:
        new = json.load(f)
    df_new = pd.DataFrame([new])
    # 3. append & drop exact duplicates
    df_out = pd.concat([df_existing, df_new], ignore_index=True)
    df_out = df_out.drop_duplicates()
    # 4. write back
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python json_to_excel.py <input.json> <output.xlsx>")
        sys.exit(1)
    append_to_excel(sys.argv[1], sys.argv[2])
